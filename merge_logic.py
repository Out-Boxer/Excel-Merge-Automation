# merge_logic.py

import openpyxl # Excel(.xlsx) 파일을 읽고 쓰는 라이브러리
from openpyxl.utils.cell import get_column_letter # 열 번호(1,2,3...)를 엑셀 열 문자('A','B',...)로 변환하는 함수를 가져옵니다
from openpyxl.utils.exceptions import InvalidFileException # 특정 예외 처리를 위해 임포트
from copy import copy
import os
import zipfile # openpyxl 라이브러리는 암호가 걸린 파일을 열 수 없기 때문에 암호가 걸린 파일에 대한 예외 처리를 하기 위한 라이브러리
import gc # 메모리 누수 방지를 위해 가바지 컬렉션을 수동으로 조작하기 위해 사용(동일 파일 병합 기준 547MB -> 520MB로 감소함)

def merge_excel_files(output_path, selected_files, gui_queue): # 인자로 필요한 변수들 GUI.py로부터 받음
    gc.disable() # 메모리 사용량 최적화를 위한 자동 가비지 컬렉션 비활성화 (메모리 수거 시점 최적화 목적)

    gui_queue.put(('log', "="*50)) 
    gui_queue.put(('log', f"병합 작업을 시작합니다. 저장 파일: {os.path.basename(output_path)}"))
    
    new_wb = None # 예외 발생 시에도 close()를 호출하기 위해 미리 선언
    try:
        new_wb = openpyxl.Workbook() # Workbook 클래스로부터 객체를 하나 생성(새 엑셀 파일을 여는 것)
        sheet_next_row = {} # 각 시트명별로 다음에 데이터를 붙여넣을 시작 행 번호를 저장(key: 시트 이름 / value: 그 시트에서 다음에 붙여넣을 행 번호)

        for file_path in selected_files: # GUI 리스트박스에서 정한 순서대로 병합
            load_file_wb = None # 파일 열기 실패해도 finally에서 안전하게 close하려고 미리 선언
            
            try: 
                gui_queue.put(('log', f"\n[{os.path.basename(file_path)}] 파일 처리 중..."))
                load_file_wb = openpyxl.load_workbook(file_path) # 선택한 파일들 불러오기

                for sheet_name in load_file_wb.sheetnames: 
                    current_sheet = load_file_wb[sheet_name] # 현재 처리 중인 시트를 current_sheet 변수에 대입
                    gui_queue.put(('log', f"  '{sheet_name}' 시트 복사 중..."))

                    if sheet_name not in new_wb.sheetnames: # 새로운 시트라면
                        gui_queue.put(('log', f"  새로운 시트 발견: '{sheet_name}' → 새 시트로 추가합니다."))
                        target_sheet = new_wb.create_sheet(title=sheet_name) # target_sheet 변수는 처리 중인 현재 시트를 가리키기 위한 변수
                        sheet_next_row[sheet_name] = 1 # 처음 발견된 시트이니까 데어터를 붙여넣을 때 1행부터 붙여넣으라는 의미
                    else: # 새로운 시트가 아니러면(이미 존재하는 시트라면)
                        target_sheet = new_wb[sheet_name]
                    start_row = sheet_next_row[sheet_name] # 시작 행 번호?

                    # row는 해당 행 한 줄의 모든 셀들의 정보(병합 상태, 현재 시트, 위치)를 저장
                    for r_idx, row in enumerate(current_sheet.iter_rows(), start=1): # openpyxl 라이브러리의 iter_rows 함수는 엑셀 시트의 행을 순회하며 데이터에 접근할 수 있게 함
                        target_r_idx = start_row + r_idx - 1
                        for c_idx, cell in enumerate(row, start=1): # cell = A1 B1 같은 셀 위치(?)
                            target_cell = target_sheet.cell(row=target_r_idx, column=c_idx) # Cell 속성을 사용하려면 행,열에 대한 위치값을 각각 알면 됩니다 / Cell( ) 내부의 숫자는 행 row=1(행), column=2(열)을 의미
                            
                            if cell.value is not None: # 셀에 실제 값이 있는 경우에만 값을 복사합니다.
                                # .value를 붙여줘야 실제값을 읽어올 수 있다는 점은 주의할 것 / .value를 붙이지 않으면 각 Cell 위치에 대한 정보값을 튜플로 읽어오기
                                target_cell.value = cell.value 
                            
                            # 원본 서식을 유지하며 셀 단위로 복사
                            if cell.has_style: # .has_style → 서식이 적용되어 있는지 여부
                                # 기존 셀 스타일(서식)을 복제할 때는 copy 함수를 사용한다
                                # 원하는 스타일만 선택해서 copy() 함수에 전달하면 해당 스타일만 카피할 수 있다
                                target_cell.font = copy(cell.font) # 현재 선택된 셀(target_cell.font)에 원본 셀의 폰트(cell.font)를 복사한다
                                target_cell.border = copy(cell.border)
                                target_cell.fill = copy(cell.fill)
                                target_cell.number_format = cell.number_format
                                target_cell.protection = copy(cell.protection)
                                target_cell.alignment = copy(cell.alignment)
                    
                    # 병합된 셀 정보 복사
                    """
                    엑셀에서 A1:B1을 병합했다고 합시다
                    엑셀 화면에서는 A1만 보이고 B1은 가려져요

                    하지만 실제로는 A1, B1 두 셀이 다 존재합니다

                    A1 → 값 있음
                    B1 → 값 없음 (빈칸처럼 보임)

                    대신 “이 구간은 병합됐다”는 별도의 병합 정보가 따로 기록됩니다
                    """
                    gui_queue.put(('log', f"    - 병합된 셀 정보 복사 중..."))
                    for merged_range in current_sheet.merged_cells.ranges: # merged_range는 A1:D1 같은 병합된 범위를 나타내는 객체                            
                        # 행 위치를 현재 데이터를 붙여넣는 위치에 맞게 조정합니다
                        # 엑셀 행 번호: 1, 2, 3, ... → 1부터 시작
                        # 코드에서 루프 돌릴 때 인덱스: 0, 1, 2, ... → 0부터 시작 / 즉, 행 위치 계산할 때 1차이를 고려해야 함

                        offset_row = start_row - 1 # start_row는 1부터 시작하므로 오프셋은 start_row - 1이 됨
                        new_min_row = merged_range.min_row + offset_row
                        new_max_row = merged_range.max_row + offset_row
                        
                        # 새로운 병합 범위를 계산
                        """
                        merged_range는 이런 속성을 갖습니다

                        min_col, min_row: 병합 범위의 시작 열/행
                        max_col, max_row: 병합 범위의 끝 열/행
                        """
                        new_range_coord = ( # 참고 튜플 아님 문자열이 저장되고 있으니 헷갈리지 말 것
                            # merged_range 객체로부터 min_col, min_row, max_col, max_row 속성을 얻을 수 있음 
                            f"{get_column_letter(merged_range.min_col)}{new_min_row}:" # 병합된 셀의 시작 위치
                            f"{get_column_letter(merged_range.max_col)}{new_max_row}" # 병합된 셀의 마지막 위치
                        )
                        target_sheet.merge_cells(new_range_coord) # 대상 시트에 셀 병합을 적용합니다. (merge_cells 병합 기능을 가진 함수)

                    # 열 너비 복사 (원본 코드의 개선된 로직 유지)
                    for col in range(1, current_sheet.max_column + 1):
                        col_letter = get_column_letter(col)
                        new_width = current_sheet.column_dimensions[col_letter].width
                        current_width = target_sheet.column_dimensions[col_letter].width
                        
                        if new_width is not None:
                            if current_width is None or new_width > current_width:
                                    target_sheet.column_dimensions[col_letter].width = new_width # column_dimensions은 열의 넓이를 지정하는 함수

                    sheet_next_row[sheet_name] += current_sheet.max_row # 다음에 데이터를 추가할 위치 업데이트
            
            finally: # <<<--- 수정된 부분: 개별 파일 처리를 위한 finally 블록
                if load_file_wb:
                    load_file_wb.close() # 메모리 관리를 위해 사용이 끝난 워크북을 명시적으로, 그리고 반드시 닫습니다.
                    gui_queue.put(('log', f"    - '{os.path.basename(file_path)}' 파일 처리 완료 및 리소스 해제"))

                    del load_file_wb
                    gc.collect() # 파일 하나가 끝날 때마다 수동으로 GC 호출(메모리 누수 방지)
        
        if 'Sheet' in new_wb.sheetnames and new_wb['Sheet'].max_row == 1 and new_wb['Sheet'].max_column == 1:
            if len(new_wb.sheetnames) > 1:
                del new_wb['Sheet']

        new_wb.save(output_path)
        gui_queue.put(('log', "\n" + "="*50))
        gui_queue.put(('log', f"성공! 모든 파일이 '{os.path.basename(output_path)}'으로 병합되었습니다."))
        gui_queue.put(('show_info', ("성공", "파일 병합이 완료되었습니다!")))

    except InvalidFileException: # 사용자가 엑셀 파일이 아닌 것을 선택했거나 파일이 손상된 경우
        user_msg = "엑셀 파일 형식이 아니거나 손상된 파일이 포함되어 있습니다.\n\n선택한 파일 목록을 확인해주세요."
        gui_queue.put(('log', f"\n오류 발생: {user_msg.replace('\n\n', ' ')}"))
        gui_queue.put(('show_error', ("파일 형식 오류", user_msg)))

    except PermissionError: # 파일 읽기/쓰기 권한이 없는 경우 (예: 파일이 다른 곳에서 열려 있음)
        user_msg = "파일에 접근할 권한이 없습니다.\n\n병합할 파일이나 저장하려는 파일이 다른 프로그램에서 열려 있는지 확인해주세요."
        gui_queue.put(('log', f"\n오류 발생: {user_msg.replace('\n\n', ' ')}"))
        gui_queue.put(('show_error', ("권한 오류", user_msg)))

    except zipfile.BadZipFile: # openpyxl는 비밀번호가 걸린 파일을 열 수 없음 / 이 에러는 주로 비밀번호가 걸린 파일을 열려고 하면 발생
        user_msg = "엑셀 파일에 비밀번호가 걸려 있으면 파일을 열 수 없습니다.\n\n비밀번호를 제거(해제 X) 후 다시 시도해주세요."
        gui_queue.put(('log', f"\n오류 발생: {user_msg.replace('\n\n', ' ')}"))
        gui_queue.put(('show_error', ("권한 오류", user_msg)))

    except Exception: # 'as e'를 사용하지 않아 예외 객체의 민감한 내용 노출을 원천 차단
        user_msg = "알 수 없는 오류가 발생하여 작업을 완료할 수 없습니다.\n\n프로그램을 다시 시작하거나 파일에 문제가 없는지 확인해주세요."
        gui_queue.put(('log', f"\n오류 발생: 예상치 못한 문제가 발생했습니다."))
        gui_queue.put(('show_error', ("오류 발생", user_msg)))

    finally: 
        if new_wb:
            new_wb.close()
            del new_wb

        gc.collect()
        gc.enable() # 자동 가비지 컬렉션 활성화
        gui_queue.put(('task_done', None))

    """
    파이썬 메모리에서 객체를 지우고 싶을 때는 del 키워드와 gc.collect()을 같이 사용해야 함
    del 키워드만 사용해서 객체를 지울 경우 네임스페이스에서 해당 객체가 사라져 참조는 안되지만
    여전히 메모리 상에는 객체가 존재하는 상태이기 때문에 꼭 gc.collect()를 같이 해주어야 한다고 함
    """